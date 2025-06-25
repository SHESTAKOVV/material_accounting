import datetime

import openpyxl
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from .forms import CustomLoginForm, MaterialForm, DirectionForm, LocationForm, SupplierForm, MaterialIncomeForm, \
    IncomeItemForm, IncomeItemFormSet, MaterialTransferForm, TransferItemFormSet, MaterialWriteOffForm, \
    WriteOffItemFormSet
from .models import Material, Direction, Location, Supplier, MaterialIncome, MaterialTransfer, MaterialWriteOff, Stock, \
    IncomeItem
from django.contrib import messages


def home(request):
  return render(request, 'home.html', {})



def logout_action(request):
    if request.user.is_authenticated:
        logout(request)

    return redirect('home', permanent=False)


def custom_login(request):
    error = None

    if request.method == "POST":
        form = CustomLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            password = form.cleaned_data["password"]
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect("home")
            else:
                error = "Неверное имя пользователя или пароль."
    else:
        form = CustomLoginForm()

    return render(request, "custom_login.html", {"form": form, "error": error})


@login_required
def material_list(request):
    query = request.GET.get("q")
    materials = Material.objects.all()
    if query:
        materials = materials.filter(name__icontains=query)
    return render(request, "material_list.html", {"materials": materials})


@login_required
def material_create(request):
    if request.method == "POST":
        form = MaterialForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("material_list")
    else:
        form = MaterialForm()
    return render(request, "material_form.html", {"form": form, "title": "Добавить материал"})


@login_required
def material_edit(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == "POST":
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            return redirect("material_list")
    else:
        form = MaterialForm(instance=material)
    return render(request, "material_form.html", {"form": form, "title": "Редактировать материал"})


@login_required
def material_delete(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == "POST":
        material.delete()
        return redirect("material_list")
    return render(request, "material_confirm_delete.html", {"material": material})


@login_required
def direction_list(request):
    directions = Direction.objects.all()
    return render(request, "direction_list.html", {"directions": directions})

@login_required
def direction_create(request):
    form = DirectionForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect("direction_list")
    return render(request, "simple_form.html", {"form": form, "title": "Добавить направление"})

@login_required
def direction_edit(request, pk):
    obj = get_object_or_404(Direction, pk=pk)
    form = DirectionForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        return redirect("direction_list")
    return render(request, "simple_form.html", {"form": form, "title": "Редактировать направление"})

@login_required
def direction_delete(request, pk):
    obj = get_object_or_404(Direction, pk=pk)
    if request.method == "POST":
        obj.delete()
        return redirect("direction_list")
    return render(request, "confirm_delete.html", {"object": obj, "title": "Удалить направление"})

# --- Location ---
@login_required
def location_list(request):
    locations = Location.objects.all()
    return render(request, "location_list.html", {"locations": locations})


@login_required
def location_create(request):
    form = LocationForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect("location_list")
    return render(request, "simple_form.html", {"form": form, "title": "Добавить место хранения"})


@login_required
def location_edit(request, pk):
    obj = get_object_or_404(Location, pk=pk)
    form = LocationForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        return redirect("location_list")
    return render(request, "simple_form.html", {"form": form, "title": "Редактировать место хранения"})


@login_required
def location_delete(request, pk):
    obj = get_object_or_404(Location, pk=pk)
    if request.method == "POST":
        obj.delete()
        return redirect("location_list")
    return render(request, "confirm_delete.html", {"object": obj, "title": "Удалить место хранения"})


# --- Supplier ---
@login_required
def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, "supplier_list.html", {"suppliers": suppliers})


@login_required
def supplier_create(request):
    form = SupplierForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect("supplier_list")
    return render(request, "simple_form.html", {"form": form, "title": "Добавить поставщика"})


@login_required
def supplier_edit(request, pk):
    obj = get_object_or_404(Supplier, pk=pk)
    form = SupplierForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        return redirect("supplier_list")
    return render(request, "simple_form.html", {"form": form, "title": "Редактировать поставщика"})


@login_required
def supplier_delete(request, pk):
    obj = get_object_or_404(Supplier, pk=pk)
    if request.method == "POST":
        obj.delete()
        return redirect("supplier_list")
    return render(request, "confirm_delete.html", {"object": obj, "title": "Удалить поставщика"})


@login_required
def income_list(request):
    incomes = MaterialIncome.objects.all().order_by("-date")
    return render(request, "income_list.html", {"incomes": incomes})

@login_required
def income_detail(request, pk):
    income = get_object_or_404(MaterialIncome, pk=pk)
    return render(request, "income_detail.html", {"income": income})

@login_required
def income_delete(request, pk):
    income = get_object_or_404(MaterialIncome, pk=pk)
    if request.method == "POST":
        income.delete()
        return redirect("income_list")
    return render(request, "confirm_delete.html", {"object": income, "title": "Удалить поступление"})


@login_required
def income_create(request):
    if request.method == "POST":
        form_income = MaterialIncomeForm(request.POST)
        formset = IncomeItemFormSet(request.POST)
        if form_income.is_valid() and formset.is_valid():
            income = form_income.save(commit=False)
            income.responsible = request.user
            income.save()
            items = formset.save(commit=False)
            for item in items:
                item.income = income
                item.save()
            return redirect("income_list")
    else:
        form_income = MaterialIncomeForm()
        formset = IncomeItemFormSet()

    return render(request, "income_formset.html", {
        "form_income": form_income,
        "formset": formset,
        "title": "Добавить поступление"
    })

@login_required
def transfer_list(request):
    transfers = MaterialTransfer.objects.all().order_by("-date")
    return render(request, "transfer_list.html", {"transfers": transfers})


@login_required
def transfer_create(request):
    if request.method == "POST":
        form_transfer = MaterialTransferForm(request.POST)
        formset = TransferItemFormSet(request.POST)
        if form_transfer.is_valid() and formset.is_valid():
            transfer = form_transfer.save(commit=False)
            transfer.responsible = request.user
            transfer.save()
            items = formset.save(commit=False)
            for item in items:
                item.transfer = transfer
                item.save()
            return redirect("transfer_list")
    else:
        form_transfer = MaterialTransferForm()
        formset = TransferItemFormSet()

    return render(request, "transfer_formset.html", {
        "form_transfer": form_transfer,
        "formset": formset,
        "title": "Добавить перемещение"
    })


@login_required
def transfer_detail(request, pk):
    transfer = get_object_or_404(MaterialTransfer, pk=pk)
    return render(request, "transfer_detail.html", {"transfer": transfer})


@login_required
def transfer_delete(request, pk):
    transfer = get_object_or_404(MaterialTransfer, pk=pk)
    if request.method == "POST":
        transfer.delete()
        return redirect("transfer_list")
    return render(request, "confirm_delete.html", {"object": transfer, "title": "Удалить перемещение"})


@login_required
def writeoff_list(request):
    writeoffs = MaterialWriteOff.objects.all().order_by("-date")
    return render(request, "writeoff_list.html", {"writeoffs": writeoffs})


@login_required
def writeoff_create(request):
    if request.method == "POST":
        form_writeoff = MaterialWriteOffForm(request.POST)
        formset = WriteOffItemFormSet(request.POST)
        if form_writeoff.is_valid() and formset.is_valid():
            writeoff = form_writeoff.save(commit=False)
            writeoff.responsible = request.user
            writeoff.save()
            items = formset.save(commit=False)
            for item in items:
                item.writeoff = writeoff
                item.save()
            return redirect("writeoff_list")
    else:
        form_writeoff = MaterialWriteOffForm()
        formset = WriteOffItemFormSet()

    return render(request, "writeoff_formset.html", {
        "form_writeoff": form_writeoff,
        "formset": formset,
        "title": "Списание материалов"
    })


@login_required
def writeoff_detail(request, pk):
    writeoff = get_object_or_404(MaterialWriteOff, pk=pk)
    return render(request, "writeoff_detail.html", {"writeoff": writeoff})


@login_required
def writeoff_delete(request, pk):
    writeoff = get_object_or_404(MaterialWriteOff, pk=pk)
    if request.method == "POST":
        writeoff.delete()
        return redirect("writeoff_list")
    return render(request, "confirm_delete.html", {"object": writeoff, "title": "Удалить списание"})


@login_required
def report_stock(request):
    stocks = Stock.objects.select_related("material", "direction", "location").order_by("material__name")

    query = request.GET.get("q")
    if query:
        stocks = stocks.filter(material__name__icontains=query)

    return render(request, "report_stock.html", {"stocks": stocks})


@login_required
def report_movement(request):
    start = request.GET.get("start")
    end = request.GET.get("end")

    incomes = MaterialIncome.objects.all()
    transfers = MaterialTransfer.objects.all()
    writeoffs = MaterialWriteOff.objects.all()

    if start and end:
        start_date = datetime.strptime(start, "%Y-%m-%d").date()
        end_date = datetime.strptime(end, "%Y-%m-%d").date()
        incomes = incomes.filter(date__range=(start_date, end_date))
        transfers = transfers.filter(date__range=(start_date, end_date))
        writeoffs = writeoffs.filter(date__range=(start_date, end_date))

    return render(request, "report_movement.html", {
        "incomes": incomes,
        "transfers": transfers,
        "writeoffs": writeoffs,
        "start": start,
        "end": end,
    })


@login_required
def report_deficit(request):
    threshold = 10  # можно сделать настраиваемым
    deficit = Stock.objects.filter(quantity__lt=threshold).select_related("material", "location", "direction")

    return render(request, "report_deficit.html", {"deficit": deficit, "threshold": threshold})


@login_required
def export_movement_excel(request):
    start = request.GET.get("start")
    end = request.GET.get("end")

    # Дата фильтрация
    incomes = MaterialIncome.objects.all()
    transfers = MaterialTransfer.objects.all()
    writeoffs = MaterialWriteOff.objects.all()

    if start and end:
        start_date = datetime.strptime(start, "%Y-%m-%d").date()
        end_date = datetime.strptime(end, "%Y-%m-%d").date()
        incomes = incomes.filter(date__range=(start_date, end_date))
        transfers = transfers.filter(date__range=(start_date, end_date))
        writeoffs = writeoffs.filter(date__range=(start_date, end_date))

    # Создание Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Движение"

    # Заголовок
    ws.append(["Тип", "Дата", "Описание", "Материал", "Кол-во", "Склад/Направление"])

    # Поступления
    for income in incomes:
        for item in income.items.all():
            ws.append([
                "Поступление",
                income.date,
                f"Поставщик: {income.supplier}",
                item.material.name,
                float(item.quantity),
                f"{item.location} / {item.direction}",
            ])

    # Перемещения
    for tr in transfers:
        for item in tr.items.all():
            ws.append([
                "Перемещение",
                tr.date,
                f"{item.from_location} → {item.to_location}",
                item.material.name,
                float(item.quantity),
                f"{item.from_direction} → {item.to_direction}",
            ])

    # Списания
    for wr in writeoffs:
        for item in wr.items.all():
            ws.append([
                "Списание",
                wr.date,
                wr.reason,
                item.material.name,
                float(item.quantity),
                f"{item.location} / {item.direction}",
            ])

    # Отправка
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=movement_report.xlsx'
    wb.save(response)
    return response


@login_required
def export_deficit_excel(request):
    threshold = 10  # фиксированный порог (можно сделать настройку)
    deficit = Stock.objects.filter(quantity__lt=threshold).select_related("material", "location", "direction")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дефицит"

    # Заголовки
    ws.append(["Материал", "Артикул", "Остаток", "Ед. изм.", "Склад", "Направление"])

    for item in deficit:
        ws.append([
            item.material.name,
            item.material.article,
            float(item.quantity),
            item.material.unit.name,
            item.location.name,
            item.direction.name
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=deficit_report.xlsx'
    wb.save(response)
    return response


@login_required
def import_income_excel(request):
    if request.method == "POST" and request.FILES.get("file"):
        wb = openpyxl.load_workbook(request.FILES["file"])
        ws = wb.active

        current_user = request.user
        created_docs = 0
        error_rows = []

        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                date = row[0].value
                supplier_name = row[1].value
                material_name = row[2].value
                quantity = row[3].value
                direction_name = row[4].value
                location_name = row[5].value

                # Поиск или создание справочников
                supplier, _ = Supplier.objects.get_or_create(name=supplier_name)
                material = Material.objects.get(name=material_name)
                direction = Direction.objects.get(name=direction_name)
                location = Location.objects.get(name=location_name)

                # Группировка: каждый импорт — отдельный документ
                income = MaterialIncome.objects.create(
                    date=date,
                    supplier=supplier,
                    responsible=current_user
                )

                IncomeItem.objects.create(
                    income=income,
                    material=material,
                    quantity=quantity,
                    direction=direction,
                    location=location
                )
                created_docs += 1
            except Exception as e:
                error_rows.append((i, str(e)))

        if error_rows:
            for row_num, err in error_rows:
                messages.error(request, f"Ошибка в строке {row_num}: {err}")
        else:
            messages.success(request, f"Успешно импортировано {created_docs} поступлений.")

        return redirect("income_list")

    return render(request, "import_income.html")